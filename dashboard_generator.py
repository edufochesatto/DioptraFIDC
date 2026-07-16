"""
DIOPTRA FIDC — Dashboard de Governança para FIDCs de Duplicatas/PME
Autor: Eduardo Fochesatto
Licença: MIT
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils.dataframe import dataframe_to_rows

# ── CONFIGURAÇÃO ──
NOME_FERRAMENTA = "Dioptra FIDC"
AUTOR = "Eduardo Fochesatto"
REPOSITORIO = "https://github.com/eduardofochesatto/dioptra-fidc"
VERSAO = "Julho/2026"

# ── Cores ──
AZUL_ESCURO = "1F3864"
BRANCO = "FFFFFF"
VERDE = "C6EFCE"
VERMELHO = "FFC7CE"
AMARELO = "FFEB9C"
CINZA_L = "F5F5F5"
CINZA_T = "808080"
PRETO = "1A1A1A"

FILL_H = PatternFill("solid", fgColor=AZUL_ESCURO)
FILL_Z = PatternFill("solid", fgColor=CINZA_L)
FILL_V = PatternFill("solid", fgColor=VERDE)
FILL_R = PatternFill("solid", fgColor=VERMELHO)
FILL_A = PatternFill("solid", fgColor=AMARELO)

FONT_H = Font(name="Calibri", bold=True, color=BRANCO, size=11)
FONT_T = Font(name="Calibri", bold=True, size=18, color=AZUL_ESCURO)
FONT_T2 = Font(name="Calibri", bold=True, size=14, color=AZUL_ESCURO)
FONT_S = Font(name="Calibri", italic=True, size=11, color=CINZA_T)
FONT_D = Font(name="Calibri", bold=True, size=12, color=AZUL_ESCURO)
FONT_N = Font(name="Calibri", size=10)
FONT_A = Font(name="Calibri", bold=True, size=10, color="CC0000")

AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")

BENCH = {
    "CDI": {"2022": 12.40, "2023": 13.05, "2024": 10.80, "2025": 11.20, "2026": 10.50},
    "IMAB": {"2022": 10.10, "2023": 15.80, "2024": 8.40, "2025": 12.50, "2026": 11.80},
    "IPCA": {"2022": 5.79, "2023": 4.62, "2024": 4.30, "2025": 4.80, "2026": 4.50},
}

def _cab(ws, l, n):
    for c in range(1, n + 1):
        cl = ws.cell(row=l, column=c)
        cl.fill = FILL_H
        cl.font = FONT_H
        cl.alignment = AC

def _zeb(ws, i, f, n):
    for r in range(i, f + 1):
        if (r - i) % 2 == 1:
            for c in range(1, n + 1):
                ws.cell(row=r, column=c).fill = FILL_Z

def _assin(ws, l):
    ws.cell(row=l, column=1,
            value=f"{NOME_FERRAMENTA} — {AUTOR} — {REPOSITORIO}").font = FONT_S

def _ajustar_largura(ws, dx):
    from openpyxl.utils import get_column_letter
    for ci, col in enumerate(dx.columns, 1):
        comp = max(len(str(col)),
                   dx[col].astype(str).str.len().max() if len(dx) > 0 else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(comp + 3, 35)
def criar_capa(wb):
    ws = wb.active
    ws.title = "0. Capa"
    itens = [
        (1, f"{NOME_FERRAMENTA}", FONT_T),
        (2, "Dashboard de Governança para FIDCs de Duplicatas/PME", FONT_D),
        (3, f"{AUTOR} — {VERSAO}", FONT_S),
        (5, "╔╗", None),
        (6, "║           AVISO IMPORTANTE                       ║", None),
        (7, "╠╣", None),
        (8,  "║  Dados públicos autodeclarados à CVM.           ║", None),
        (9,  "║  LIMITAÇÕES: defasagem 30-45d, preenchimento    ║", None),
        (10, "║  parcial, classificação aproximada.             ║", None),
        (11, "╠╣", None),
        (12, "║  NÃO SUBSTITUI DILIGÊNCIA — triagem, não decisão║", None),
        (13, "╚╝", None),
        (15, f"📬 {AUTOR}", FONT_D),
        (16, f"Repositório: {REPOSITORIO}", FONT_S),
        (17, "Licença: MIT", FONT_S),
        (19, "📋 Abas: 0.Capa | 1.Ranking | 2.Top10 | 3.CDI vs IMAB | 4.Governança | 5.Propósito | 6.Fontes | 7.Glossário", FONT_N),
    ]
    for l, txt, st in itens:
        c = ws.cell(row=l, column=1, value=txt)
        if st:
            c.font = st
    for l in [6, 10, 12]:
        ws.cell(row=l, column=1).font = Font(name="Consolas", bold=True, size=10, color="CC0000")
    for l in range(5, 14):
        ws.cell(row=l, column=1).font = Font(name="Consolas", size=10, color=PRETO)
    ws.column_dimensions["A"].width = 100
    ws.sheet_properties.tabColor = "CC0000"

def criar_ranking(wb, df):
    """Aba 1: Ranking completo com filtro automático."""
    ws = wb.create_sheet("1. Ranking Completo")

    MAPA = {
        'CNPJ_FUNDO': 'CNPJ do Fundo',
        'DENOMINACAO_SOCIAL': 'Nome do Fundo',
        'GESTORA': 'Gestora',
        'CLASSE': 'Classe',
        'VL_PL': 'PL (R$ milhões)',
        'PRAZO_MEDIO': 'Prazo Médio (dias)',
        'RENTABILIDADE': 'Rentabilidade (% a.a.)',
        'PDD_PCT': 'PDD (% PL)',
        'RECOMPRA_PCT': 'Recompra (%)',
        'PCT_SUBORDINADO': 'Subordinado (%)',
        'OVERCOLLATERALIZATION': 'Overcoll. (x)',
        'NUM_SACADOS': 'Nº Sacados',
        'NUM_CEDENTES': 'Nº Cedentes',
        'CONC_TOP5_SACADOS': 'Conc. Top5 Sacados (%)',
        'CONC_TOP5_CEDENTES': 'Conc. Top5 Cedentes (%)',
        'PCT_VENCIDOS': 'Vencidos +90d (%)',
        'PCT_LIQUIDEZ': 'Liquidez (%)',
        'TAXA_ADM': 'Taxa Adm (% a.a.)',
        'RATING': 'Rating',
    }
    cols = [c for c in MAPA if c in df.columns]
    if not cols:
        print("   [AVISO] Nenhuma coluna do mapa encontrada para o ranking.")
        ws.cell(row=1, column=1, value="Nenhum dado disponível para exibir.").font = FONT_D
        return
    dx = df[cols].copy()
    dx.columns = [MAPA[c] for c in cols]

    if 'PL (R$ milhões)' in dx.columns:
        dx['PL (R$ milhões)'] = pd.to_numeric(dx['PL (R$ milhões)'], errors='coerce').fillna(0)
        dx['PL (R$ milhões)'] = (dx['PL (R$ milhões)'] / 1_000_000).round(2)
        dx = dx.sort_values('PL (R$ milhões)', ascending=False)
    elif len(dx.columns) > 1:
        for c in dx.columns:
            if dx[c].dtype in ('float64', 'int64'):
                dx = dx.sort_values(c, ascending=False)
                break
        else:
            dx = dx.sort_values(dx.columns[0], ascending=False)

    for r, row in enumerate(dataframe_to_rows(dx, index=False, header=True), 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)

    nc = len(dx.columns)
    if nc == 0:
        return
    ul = len(dx) + 1
    _cab(ws, 1, nc)
    _zeb(ws, 2, ul, nc)
    _ajustar_largura(ws, dx)
    _assin(ws, ul + 2)
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=nc).column_letter}{ul}"
    ws.freeze_panes = "A2"

def criar_top10(wb, df):
    """Aba 2: Top 10 fundos em cada métrica."""
    ws = wb.create_sheet("2. Top 10 por Métrica")

    METS = [
        ("PL (R$ milhões)", "VL_PL", False),
        ("Rentabilidade (% a.a.)", "RENTABILIDADE", False),
        ("Menor PDD (% PL)", "PDD_PCT", True),
        ("Maior Overcollateralization", "OVERCOLLATERALIZATION", False),
        ("Maior Liquidez (%)", "PCT_LIQUIDEZ", False),
        ("Menor Conc. Top5 Sacados (%)", "CONC_TOP5_SACADOS", True),
        ("Menor Conc. Top5 Cedentes (%)", "CONC_TOP5_CEDENTES", True),
        ("Maior Nº de Sacados", "NUM_SACADOS", False),
        ("Menor % Vencidos +90d", "PCT_VENCIDOS", True),
        ("Maior Recompra (%)", "RECOMPRA_PCT", False),
    ]

    linha = 1
    for titulo, coluna, asc in METS:
        if coluna not in df.columns:
            continue
        v = df[coluna].dropna()
        if len(v) == 0:
            continue
        top = df.sort_values(coluna, ascending=asc).head(10)

        ws.cell(row=linha, column=1, value=f"TOP 10 — {titulo}").font = FONT_D
        linha += 1
        for c, t in enumerate(["#", "Fundo", "Gestora", "Valor"], 1):
            ws.cell(row=linha, column=c, value=t)
        _cab(ws, linha, 4)
        linha += 1

        for idx, (_, row) in enumerate(top.iterrows(), 1):
            ws.cell(row=linha, column=1, value=idx).alignment = AC
            ws.cell(row=linha, column=2, value=str(row.get('DENOMINACAO_SOCIAL', ''))[:60])
            ws.cell(row=linha, column=3, value=str(row.get('GESTORA', ''))[:30])
            val = row[coluna]
            if coluna == 'VL_PL':
                val = round(val / 1_000_000, 2)
            ws.cell(row=linha, column=4, value=round(val, 2)).alignment = AR
            linha += 1
        linha += 2

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    _assin(ws, linha + 1)
def criar_comparativo(wb, df):
    """Aba 3: Comparação com CDI e IMAB."""
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("3. CDI vs IMAB")

    ws.cell(row=1, column=1,
            value=f"{NOME_FERRAMENTA} — FIDCs vs CDI vs IMAB").font = FONT_T
    ws.cell(row=2, column=1,
            value=f"Fontes: CVM | ANBIMA Data | {AUTOR}").font = FONT_S
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    CDIa = BENCH["CDI"].get("2026", 10.50)
    IMAa = BENCH["IMAB"].get("2026", 11.80)

    linha = 4
    cab = ["Fundo", "Rentabilidade", f"CDI ({CDIa}%)", f"IMAB ({IMAa}%)",
           "vs CDI (p.p.)", "vs IMAB (p.p.)", "Rating", "Classe"]
    for c, t in enumerate(cab, 1):
        ws.cell(row=linha, column=c, value=t)
    _cab(ws, linha, len(cab))
    linha += 1

    if 'RENTABILIDADE' in df.columns:
        df_temp = df.dropna(subset=['RENTABILIDADE']).sort_values('RENTABILIDADE', ascending=False).head(20)
        for _, row in df_temp.iterrows():
            rent = row['RENTABILIDADE']
            ws.cell(row=linha, column=1, value=str(row.get('DENOMINACAO_SOCIAL', ''))[:45])
            ws.cell(row=linha, column=2, value=round(rent, 2)).alignment = AR
            ws.cell(row=linha, column=3, value=CDIa).alignment = AR
            ws.cell(row=linha, column=4, value=IMAa).alignment = AR
            vc, vi = rent - CDIa, rent - IMAa
            for col, val in [(5, vc), (6, vi)]:
                c = ws.cell(row=linha, column=col, value=round(val, 2))
                c.alignment = AR
                cor = "006100" if val > 0 else "9C0006"
                c.font = Font(name="Calibri", bold=True, color=cor)
                c.fill = FILL_V if val > 0 else FILL_R
            ws.cell(row=linha, column=7, value=row.get('RATING', ''))
            ws.cell(row=linha, column=8, value=row.get('CLASSE', ''))
            linha += 1

    linha += 3
    ws.cell(row=linha, column=1, value="Benchmarks Históricos (ANBIMA)").font = FONT_T2
    linha += 1
    cab2 = ["Ano", "CDI", "IMAB", "IPCA", "CDI Real", "IMAB Real"]
    for c, t in enumerate(cab2, 1):
        ws.cell(row=linha, column=c, value=t)
    _cab(ws, linha, 6)
    linha += 1
    for ano in sorted(BENCH["CDI"]):
        cdi = BENCH["CDI"][ano]
        imab = BENCH["IMAB"][ano]
        ipca = BENCH["IPCA"][ano]
        ws.cell(row=linha, column=1, value=ano)
        ws.cell(row=linha, column=2, value=cdi).alignment = AR
        ws.cell(row=linha, column=3, value=imab).alignment = AR
        ws.cell(row=linha, column=4, value=ipca).alignment = AR
        ws.cell(row=linha, column=5,
                value=round((1 + cdi / 100) / (1 + ipca / 100) - 1, 4)).alignment = AR
        ws.cell(row=linha, column=6,
                value=round((1 + imab / 100) / (1 + ipca / 100) - 1, 4)).alignment = AR
        linha += 1

    linha += 2
    ws.cell(row=linha, column=1,
            value="⚠ Comparação ilustrativa. Riscos distintos.").font = FONT_S
    _assin(ws, linha + 1)

    for c, w in [(1, 40), (2, 16), (3, 14), (4, 14), (5, 16), (6, 16),
                  (7, 10), (8, 16)]:
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"


def criar_governanca(wb, metricas):
    """Aba 4: Estatísticas descritivas das métricas."""
    ws = wb.create_sheet("4. Governança")
    if not metricas:
        ws.cell(row=1, column=1, value="Nenhuma métrica disponível.").font = FONT_D
        return

    MAPA = {
        'VL_PL': 'PL (R$ milhões)',
        'PRAZO_MEDIO': 'Prazo Médio (dias)',
        'PDD_PCT': 'PDD (% PL)',
        'RECOMPRA_PCT': 'Recompra (%)',
        'CONC_TOP5_SACADOS': 'Conc. Top5 Sacados (%)',
        'CONC_TOP5_CEDENTES': 'Conc. Top5 Cedentes (%)',
        'NUM_SACADOS': 'Nº Sacados',
        'NUM_CEDENTES': 'Nº Cedentes',
        'OVERCOLLATERALIZATION': 'Overcollateralization (x)',
        'PCT_LIQUIDEZ': 'Liquidez (%)',
        'PCT_VENCIDOS': 'Vencidos +90d (%)',
        'RENTABILIDADE': 'Rentabilidade (% a.a.)',
    }
    INTERP = {
        'VL_PL': 'Fundos < R$ 50 MM requerem atenção.',
        'PRAZO_MEDIO': 'Acima de 180d: risco de alongamento.',
        'PDD_PCT': 'Até 2%: ok. 2-5%: atenção. >5%: alerta.',
        'RECOMPRA_PCT': '>80%: confiança do cedente.',
        'CONC_TOP5_SACADOS': '>60%: risco severo.',
        'CONC_TOP5_CEDENTES': '>70%: dependência crítica.',
        'NUM_SACADOS': 'Mais pulverizado = menos risco.',
        'NUM_CEDENTES': 'Diversificação reduz risco.',
        'OVERCOLLATERALIZATION': '>1.2x: saudável. <1.0x: crítico.',
        'PCT_LIQUIDEZ': 'Ideal >10%. <5%: risco.',
        'PCT_VENCIDOS': 'Até 2%: normal. >5%: problema estrutural.',
        'RENTABILIDADE': 'Analisar junto com PDD e prazo médio.',
    }

    cab = ["Métrica", "Média", "Mediana", "Desv.Padrão",
           "Melhor", "Pior", "P25", "P75", "Interpretação"]
    for c, t in enumerate(cab, 1):
        ws.cell(row=1, column=c, value=t)
    _cab(ws, 1, len(cab))

    linha = 2
    for met, vals in metricas.items():
        nome = MAPA.get(met, met)
        ws.cell(row=linha, column=1, value=nome).font = Font(name="Calibri", bold=True)
        for ci, ch in enumerate(['media', 'mediana', 'desvio_padrao', 'min',
                                  'max', 'p25', 'p75'], 2):
            v = vals.get(ch)
            if v is not None:
                ws.cell(row=linha, column=ci, value=v).alignment = AR
        ws.cell(row=linha, column=9, value=INTERP.get(met, ''))

        if met in ('PDD_PCT', 'CONC_TOP5_SACADOS', 'CONC_TOP5_CEDENTES', 'PCT_VENCIDOS'):
            md = vals.get('media', 0)
            if md and md > 5:
                ws.cell(row=linha, column=2).fill = FILL_R
            elif md and md > 3:
                ws.cell(row=linha, column=2).fill = FILL_A
        linha += 1

    for c, w in [(1, 28), (2, 12), (3, 12), (4, 14), (5, 12), (6, 12),
                  (7, 10), (8, 10), (9, 70)]:
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    ws.freeze_panes = "A2"
    _assin(ws, linha + 2)

def criar_proposito(wb):
    """Aba 5: Ensaio reflexivo."""
    ws = wb.create_sheet("5. Propósito")
    textos = [
        (f"O Propósito do {NOME_FERRAMENTA}", FONT_T),
        (f"Por {AUTOR} — {VERSAO}", FONT_S),
        ("", None),
        ("O que você está prestes a ver não é uma planilha. É um mapa de minas terrestres.", FONT_D),
        ("", None),
        ("Se existe uma verdade que Morgan Housel ensinou, é esta:\n"
         "o maior risco não está no que você consegue ver — está no que você não consegue.\n\n"
         "Você olha para a rentabilidade de um FIDC e vê 13% ao ano. Lindo. Mas o que está por trás?\n"
         "Qual o prazo médio? Qual a concentração? Qual a PDD?\n\n"
         "Um fundo pode quebrar não porque rendeu pouco, mas porque ninguém olhou para o lastro.", None),
        ("", None),
        ("A 'Basileia' dos FIDCs", FONT_D),
        ("PILAR 1 — CAPITAL: Subordinação + Overcollateralization\n"
         "   → O colchão que protege o investidor sênior.\n\n"
         "PILAR 2 — RISCO: PDD + Concentração + Prazo Médio + Liquidez\n"
         "   → A gestão ativa do risco de crédito.\n\n"
         "PILAR 3 — TRANSPARÊNCIA: Informe Mensal CVM + Rating\n"
         "   → A prestação de contas pública.\n\n"
         "Diferença? Bancos são obrigados por lei a seguir Basileia. FIDCs não.\n"
         "Quando você exige esses números, você aplica padrão regulatório onde o regulador não chegou.", None),
        ("", None),
        ("⚠ A analogia tem limites:", FONT_A),
        ("• Dados autodeclarados, não auditados\n"
         "• Cada classe de FIDC tem realidades diferentes\n"
         "• Defasagem de 30-45 dias\n"
         "• Nem todo fundo preenche todos os campos\n\n"
         "Use como triagem, não como verdade absoluta.",
         Font(name="Calibri", size=10, color=CINZA_T)),
        ("", None),
        ("Como usar:", FONT_D),
        ("1. Abra o Ranking e filtre por classe\n"
         "2. Veja o Top 10 em cada métrica\n"
         "3. Compare com CDI e IMAB\n"
         "4. Interprete na aba Governança\n\n"
         "Pergunte: 'Se esse fundo estivesse no meu portfólio, eu dormiria bem?'", None),
        ("", None),
        (f"📬 {AUTOR}\n{REPOSITORIO}", FONT_S),
    ]
    linha = 1
    for txt, st in textos:
        if txt == "":
            linha += 1
            continue
        c = ws.cell(row=linha, column=1, value=txt)
        if st:
            c.font = st
        c.alignment = AL
        nl = max(len(str(txt).split('\n')), len(str(txt)) // 85 + 1)
        ws.row_dimensions[linha].height = max(nl * 16, 20)
        linha += 1
    linha += 2
    c = ws.cell(row=linha, column=1,
                value='"O segredo do bom investimento não é saber o que vai acontecer. '
                      'É aceitar que você não sabe, e se preparar para qualquer coisa."\n'
                      '— Morgan Housel, A Psicologia Financeira')
    c.font = Font(name="Calibri", italic=True, size=11, color=CINZA_T)
    c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120

def criar_fontes(wb):
    """Aba 6: Documentação das fontes."""
    ws = wb.create_sheet("6. Fontes")
    linhas = [
        ("FONTES OFICIAIS", "t2"), ("", ""),
        ("1. CVM — Portal Dados Abertos", "s"),
        ("   https://dados.cvm.gov.br/dataset/fidc-doc-inf_mensal", "l"),
        ("   Informe Mensal de FIDCs", "d"), ("", ""),
        ("2. ANBIMA Data", "s"),
        ("   https://data.anbima.com.br", "l"), ("", ""),
        ("3. Benchmarks CDI/IMAB/IPCA", "s"),
        ("   Fonte: ANBIMA / B3", "d"), ("", ""),
        ("PROCESSAMENTO", "t2"), ("", ""),
        ("  1. Acessa repositório CVM e baixa ZIP mais recente", "p"),
        ("  2. Extrai CSVs (TAB_I a TAB_IX)", "p"),
        ("  3. Filtra fundos com CLASSE = Duplicatas ou PME", "p"),
        ("  4. Calcula métricas derivadas (PDD/PL, overcoll., etc.)", "p"),
        ("  5. Gera o dashboard Excel com formatação condicional", "p"),
        ("", ""),
        ("⚠ LIMITAÇÕES", "av"), ("", ""),
        ("  • Formato dos CSVs pode mudar sem aviso", "d"),
        ("  • PDD, prazo médio e recompra frequentemente vêm zerados", "d"),
        ("  • Classificação Duplicatas/PME é aproximada", "d"),
        ("  • Defasagem de 30-45 dias", "d"), ("", ""),
        (f"{NOME_FERRAMENTA} — PÚBLICO E GRATUITO", "dest"),
        (f"Autor: {AUTOR}", ""), (f"Repositório: {REPOSITORIO}", ""),
        ("Licença: MIT", ""),
    ]
    est = {
        "t2": Font(name="Calibri", bold=True, size=14, color=AZUL_ESCURO),
        "s": Font(name="Calibri", bold=True, size=12, color="2E75B6"),
        "l": Font(name="Calibri", size=10, color="0563C1", underline="single"),
        "d": Font(name="Calibri", size=10, color=PRETO),
        "p": Font(name="Calibri", size=10),
        "av": Font(name="Calibri", bold=True, size=10, color="CC0000"),
        "dest": Font(name="Calibri", bold=True, size=11, color=AZUL_ESCURO),
    }
    for r, (txt, stl) in enumerate(linhas, 1):
        c = ws.cell(row=r, column=1, value=txt)
        if stl in est:
            c.font = est[stl]
    ws.column_dimensions["A"].width = 110

def criar_glossario(wb):
    """Aba 7: Glossário de métricas."""
    ws = wb.create_sheet("7. Glossário")
    dados = [
        ("Métrica", "O que é", "Como interpretar"),
        ("PL (R$ milhões)", "Patrimônio Líquido do fundo.",
         ">R$200 MM: líquido. <R$50 MM: atenção."),
        ("Prazo Médio (dias)", "Prazo médio ponderado dos direitos creditórios.",
         "Até 90d: curto. 90-180: médio. >180: risco."),
        ("PDD (% PL)", "Provisão para Devedores Duvidosos.",
         "Até 2%: saudável. 2-5%: atenção. >5%: alerta."),
        ("Recompra (%)", "Direitos recomprados no mês / PL.",
         ">80%: confiança. <50%: atenção."),
        ("Overcollateralization", "Ativo total / PL.",
         ">1.2x: bom. 1.0-1.2x: apertado. <1.0x: crítico."),
        ("Conc. Top5 Sacados", "% nos 5 maiores devedores.",
         "Até 40%: bom. 40-60%: moderado. >60%: severo."),
        ("Conc. Top5 Cedentes", "% nos 5 maiores cedentes.",
         "Até 50%: saudável. >70%: dependência crítica."),
        ("Liquidez (%)", "Caixa + títulos públicos / PL.",
         "Ideal >10%. <5%: risco."),
        ("Vencidos +90d (%)", "Créditos com atraso >90d / PL.",
         "Até 2%: normal. 2-5%: investigar. >5%: problema estrutural."),
        ("Subordinado (%)", "Cotas que absorvem perdas primeiro.",
         "Maior = mais proteção ao sênior. Mínimo regulatório: 10%."),
        ("CDI", "Certificado de Depósito Interbancário.",
         "Principal benchmark de renda fixa."),
        ("IMAB", "Índice ANBIMA de títulos prefixados.",
         "Benchmark para FIDCs de duration longa."),
    ]
    for r, row in enumerate(dados, 1):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    _cab(ws, 1, 3)
    _zeb(ws, 2, len(dados), 3)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 60
    ws.freeze_panes = "A2"

def gerar_dashboard(df_fundos, metricas, output_path):
    """Função principal: gera o workbook completo."""
    if df_fundos.empty:
        print("[AVISO] DataFrame vazio. Dashboard não gerado.")
        return
    wb = Workbook()
    try:
        criar_capa(wb)
        criar_ranking(wb, df_fundos)
        criar_top10(wb, df_fundos)
        criar_comparativo(wb, df_fundos)
        criar_governanca(wb, metricas)
        criar_proposito(wb)
        criar_fontes(wb)
        criar_glossario(wb)
        wb.save(output_path)
        print(f"\n✅ {NOME_FERRAMENTA} gerado com sucesso!")
        print(f"📁 {output_path}")
    except Exception as e:
        print(f"\n❌ Erro ao gerar dashboard: {e}")
        raise
